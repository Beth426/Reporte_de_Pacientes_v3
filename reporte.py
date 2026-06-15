"""
reporte.py
Lógica de generación del Reporte_Pacientes_sin_desercion.xlsx.
Importado por app.py — no tiene CLI propio.
"""

import io
import pandas as pd
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────
def _arial(**k):  return Font(name='Arial', **k)
def _fill(color): return PatternFill('solid', fgColor=color)
def _border():
    t = Side(style='thin', color='D9D9D9')
    return Border(t, t, t, t)

HDR_FILL  = _fill('1F3864')
HDR_FONT  = _arial(bold=True, color='FFFFFF', size=10)
ALT_FILLS = ['EEF1F7', 'FFFFFF']
GPOS      = _fill('E2EFDA')
GZERO     = _fill('F2F2F2')
CEN       = Alignment('center', 'center', wrap_text=True)
LEF       = Alignment('left',   'center')
BD        = _border()


# ─────────────────────────────────────────────
# NORMALIZACIÓN DE ÁREA
# ─────────────────────────────────────────────
def _normalizar_area(s):
    s = str(s).lower()
    if 'fisioterap'  in s: return 'Fisioterapia'
    if 'fonoaud' in s or 'foniatria' in s: return 'Fonoaudiología'
    if 'lenguaje'    in s: return 'Terapia del Lenguaje'
    if 'ocupacional' in s: return 'Terapia Ocupacional'
    if 'neuro'       in s: return 'Neuropsicología'
    if 'psicolog'    in s: return 'Psicología'
    if 'sensorial'   in s: return 'Integración Sensorial'
    return 'Otro'


# ─────────────────────────────────────────────
# CARGA Y LIMPIEZA
# ─────────────────────────────────────────────
def cargar_y_limpiar(archivo_bytes, hoy=None):
    """
    Recibe un objeto de bytes (BytesIO o similar) y retorna (df, n_correcciones).
    """
    if hoy is None:
        hoy = pd.Timestamp(date.today())

    df = pd.read_excel(archivo_bytes, sheet_name=0, engine='openpyxl')

    mask_c1 = df['ESTADO-ACTUAL-CITA'].astype(str) != 'FINALIZADA'
    n_corr   = int(mask_c1.sum())
    for i in df[mask_c1].index:
        df.at[i, 'SERVICIO']            = df.at[i, 'PROFESIONAL-ATIENDE']
        df.at[i, 'PROFESIONAL-ATIENDE'] = df.at[i, 'CONSENTIMIENTO-FIRMADO']
        df.at[i, 'FECHA-FIN-ATENCION']  = df.at[i, 'FECHA-CANCELA']

    df['FFIN']  = pd.to_datetime(df['FECHA-FIN-ATENCION'], errors='coerce')
    df['FDATE'] = df['FFIN'].dt.normalize()
    df['PER']   = df['FFIN'].dt.to_period('M')
    df['AREA']  = df['SERVICIO'].apply(_normalizar_area)
    df['PACIENTE'] = (df['NOMBRES'].astype(str).str.strip()
                      + ' ' + df['APELLIDOS'].astype(str).str.strip()).str.strip()
    df['PROFESIONAL-ATIENDE'] = df['PROFESIONAL-ATIENDE'].astype(str).str.strip()

    return df, n_corr, hoy


# ─────────────────────────────────────────────
# CONSTRUCCIÓN DE TABLAS
# ─────────────────────────────────────────────
def construir_tablas(df, hoy):
    MESES_ES = {1:'Enero',2:'Febrero',3:'Marzo',4:'Abril',5:'Mayo',6:'Junio',
                7:'Julio',8:'Agosto',9:'Septiembre',10:'Octubre',
                11:'Noviembre',12:'Diciembre'}
    meses_periodo = sorted(df['PER'].dropna().unique())
    nombres_mes   = {p: f"{MESES_ES[p.month]} {p.year}" for p in meses_periodo}
    meses         = [nombres_mes[p] for p in meses_periodo]

    # Reporte General
    rg = []
    for doc, g in df.groupby('DOCUMENTO'):
        g    = g.sort_values('FFIN')
        last = g.iloc[-1]
        ult  = g['FDATE'].max()
        rg.append(dict(
            Documento   = doc,
            Paciente    = g['PACIENTE'].mode().iloc[0],
            Primera     = g['FDATE'].min().date(),
            Ultima      = ult.date(),
            Area        = last['AREA'],
            Servicio    = last['SERVICIO'],
            Profesional = last['PROFESIONAL-ATIENDE'],
            Total       = len(g),
            Dias        = (hoy - ult).days,
        ))
    gen = pd.DataFrame(rg).sort_values('Paciente')

    # Reporte Mensual
    ct = (df.pivot_table(index=['DOCUMENTO','PACIENTE'], columns='PER',
                         values='FDATE', aggfunc='count', fill_value=0)
            .reset_index())
    ct.columns = ['Documento','Paciente'] + [nombres_mes[c] for c in ct.columns[2:]]
    for m in meses:
        if m not in ct.columns:
            ct[m] = 0
    ct = ct[['Documento','Paciente'] + meses]
    ct['Total'] = ct[meses].sum(axis=1)
    ct = ct.sort_values('Paciente')

    # Reporte por Profesional
    tab = (df.groupby(['PROFESIONAL-ATIENDE','AREA','DOCUMENTO','PACIENTE'])
             .agg(Evoluciones=('FFIN','count'),
                  Ultima=('FDATE','max'),
                  Primera=('FDATE','min'))
             .reset_index()
             .rename(columns={'PROFESIONAL-ATIENDE':'Profesional',
                              'AREA':'Área','DOCUMENTO':'Documento','PACIENTE':'Paciente'}))
    tab['Ultima']  = tab['Ultima'].apply(lambda x: x.date())
    tab['Primera'] = tab['Primera'].apply(lambda x: x.date())
    tab['Dias']    = tab['Ultima'].apply(lambda x: (hoy.date() - x).days)
    tab = tab.sort_values(['Profesional','Área','Paciente']).reset_index(drop=True)

    return gen, ct, tab, meses


# ─────────────────────────────────────────────
# HELPERS DE HOJAS
# ─────────────────────────────────────────────
def _encabezado(ws, headers):
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h)
        c.fill = HDR_FILL; c.font = HDR_FONT
        c.alignment = CEN; c.border = BD
    ws.row_dimensions[1].height = 28
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row or 1}"

def _ancho(ws, widths):
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = w


# ─────────────────────────────────────────────
# HOJAS
# ─────────────────────────────────────────────
def hoja_procesos(wb, df, hoy, meses):
    ws = wb.active
    ws.title = 'Procesos y Parámetros'
    ws.sheet_view.showGridLines = False
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 80

    r = 2
    ws.cell(r,2,'Centro Terapéutico — Metodología del reporte').font = \
        _arial(bold=True, size=15, color='1F3864'); r += 1
    ws.cell(r,2,'Versión SIN condición de deserción (solo actividad real)').font = \
        _arial(italic=True, size=10, color='808080'); r += 2

    def section(title):
        nonlocal r
        c = ws.cell(r, 2, title)
        c.font = _arial(bold=True, color='FFFFFF', size=11)
        c.fill = HDR_FILL; c.alignment = LEF
        ws.cell(r, 3).fill = HDR_FILL
        ws.row_dimensions[r].height = 20; r += 1

    def kv(k, v):
        nonlocal r
        ws.cell(r, 2, k).font = _arial(bold=True, size=9, color='1F3864')
        c = ws.cell(r, 3, v); c.font = _arial(size=9)
        c.alignment = Alignment('left','top', wrap_text=True)
        ws.row_dimensions[r].height = max(15, len(str(v))//6)
        r += 1

    periodo = f"{meses[0]} – {meses[-1]}" if meses else "Sin datos"
    section('1. Fuente de datos')
    kv('Archivo origen',       'semestral.xlsx — hoja "Centro Terapéutico_PACIENTES"')
    kv('Registros procesados', f'{len(df):,} evoluciones')
    kv('Período detectado',    periodo)
    kv('Pacientes únicos',     str(df["DOCUMENTO"].nunique()))
    kv('Profesionales únicos', str(df["PROFESIONAL-ATIENDE"].nunique()))
    r += 1

    section('2. Fecha usada para la evolución')
    kv('Campo',  'FECHA-FIN-ATENCION')
    kv('Motivo', 'Representa el cierre/registro real de cada evolución.')
    r += 1

    section('3. Fecha de referencia')
    kv('Fecha de corte', str(hoy.date()))
    kv('Uso', '"Días desde última evolución" = fecha de corte − última evolución del paciente.')
    r += 1

    section('4. Limpieza de datos')
    kv('Corrección 1 (automática)',
       'Filas donde ESTADO-ACTUAL-CITA ≠ "FINALIZADA": '
       'SERVICIO ← PROFESIONAL-ATIENDE, PROFESIONAL-ATIENDE ← CONSENTIMIENTO-FIRMADO, '
       'FECHA-FIN-ATENCION ← FECHA-CANCELA.')
    kv('Corrección 2 (manual)', 'El usuario corrige sufijos corridos en PROFESIONAL-ATIENDE en el archivo fuente.')
    r += 1

    section('5. Normalización de Área')
    for kw, area in [('fisioterap →','Fisioterapia'),('fonoaud / foniatria →','Fonoaudiología'),
                     ('lenguaje →','Terapia del Lenguaje'),('ocupacional →','Terapia Ocupacional'),
                     ('neuro →','Neuropsicología'),('psicolog →','Psicología'),
                     ('sensorial →','Integración Sensorial')]:
        kv(f'  {kw}', area)
    r += 1

    section('6. Contenido de hojas')
    kv('Reporte General',      'Una fila por paciente con primera/última evolución, área, profesional, total y días.')
    kv('Reporte Mensual',      'Matriz paciente × mes. Verde = actividad, gris = sin actividad.')
    kv('Reporte por Profesional', 'Profesional × Área × Paciente con conteo y fechas.')
    kv('Base de Datos',        'Copia exacta del archivo fuente sin transformaciones.')


def hoja_general(wb, gen, hoy):
    ws  = wb.create_sheet('Reporte General')
    H   = ['Documento','Paciente','Primera evolución','Última evolución',
           'Área','Servicio (exacto)','Profesional','Total evoluciones',
           f'Días desde última evolución (al {hoy.strftime("%d-%b-%Y")})']
    cols= ['Documento','Paciente','Primera','Ultima','Area','Servicio','Profesional','Total','Dias']
    _encabezado(ws, H)
    for i, (_, row) in enumerate(gen.iterrows(), 2):
        for j, col in enumerate(cols, 1):
            c = ws.cell(i, j, row[col]); c.border = BD; c.font = _arial(size=9)
            c.alignment = LEF if col in ('Paciente','Area','Servicio','Profesional') else CEN
        ws.cell(i, 8).font = _arial(size=9, bold=True)
        ws.cell(i, 9).font = _arial(size=9, bold=True)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(H))}{len(gen)+1}"
    _ancho(ws, [12,30,16,16,18,40,30,15,22])


def hoja_mensual(wb, ct, meses):
    ws = wb.create_sheet('Reporte Mensual')
    H  = ['Documento','Paciente'] + meses + ['Total']
    _encabezado(ws, H)
    for i, (_, row) in enumerate(ct.iterrows(), 2):
        for j, col in enumerate(ct.columns, 1):
            c = ws.cell(i, j, row[col]); c.border = BD; c.font = _arial(size=9)
            c.alignment = LEF if j == 2 else CEN
        for col in range(3, 3+len(meses)):
            c = ws.cell(i, col)
            c.fill = GPOS if c.value else GZERO
            if not c.value:
                c.font = _arial(size=9, color='BFBFBF')
        ws.cell(i, 3+len(meses)).font = _arial(size=9, bold=True)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(H))}{len(ct)+1}"
    _ancho(ws, [12,30] + [11]*len(meses) + [10])


def hoja_profesional(wb, tab, hoy):
    ws   = wb.create_sheet('Reporte por Profesional')
    H    = ['Profesional','Área','Documento','Paciente','Evoluciones',
            'Primera evolución','Última evolución',
            f'Días desde última evolución (al {hoy.strftime("%d-%b-%Y")})']
    cols = ['Profesional','Área','Documento','Paciente','Evoluciones','Primera','Ultima','Dias']
    _encabezado(ws, H)
    profs = tab['Profesional'].unique()
    prof_color = {p: ALT_FILLS[i % 2] for i, p in enumerate(profs)}
    for i, (_, row) in enumerate(tab.iterrows(), 2):
        FILL = _fill(prof_color[row['Profesional']])
        for j, col in enumerate(cols, 1):
            c = ws.cell(i, j, row[col])
            c.border = BD; c.fill = FILL
            c.font   = _arial(size=9, bold=(col in ('Evoluciones','Dias')))
            c.alignment = LEF if col in ('Profesional','Área','Paciente') else CEN
    ws.auto_filter.ref = f"A1:{get_column_letter(len(H))}{len(tab)+1}"
    _ancho(ws, [34,22,12,30,13,17,17,22])


def hoja_base_datos(wb, archivo_bytes):
    """Copia el archivo fuente tal cual, sin transformaciones."""
    df_raw = pd.read_excel(archivo_bytes, sheet_name=0, engine='openpyxl')
    ws = wb.create_sheet('Base de Datos')
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = 'A2'
    headers = list(df_raw.columns)
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h)
        c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = CEN; c.border = BD
    ws.row_dimensions[1].height = 28
    for i, (_, row) in enumerate(df_raw.iterrows(), 2):
        fill = _fill(ALT_FILLS[i % 2])
        for j, val in enumerate(row, 1):
            c = ws.cell(i, j, val if pd.notna(val) else None)
            c.font = _arial(size=9); c.border = BD
            c.fill = fill; c.alignment = CEN
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(df_raw)+1}"
    for j, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(j)].width = max(10, min(30, len(str(h))+4))


# ─────────────────────────────────────────────
# FUNCIÓN PRINCIPAL — retorna bytes del xlsx
# ─────────────────────────────────────────────
def generar_reporte_bytes(archivo_bytes, hoy=None):
    """
    Recibe los bytes del archivo fuente.
    Retorna (bytes_del_xlsx, resumen_dict).
    """
    # Necesitamos leer dos veces: cargar_y_limpiar y hoja_base_datos
    # Convertimos a BytesIO para poder rebobinar
    raw = io.BytesIO(archivo_bytes if isinstance(archivo_bytes, bytes)
                     else archivo_bytes.read())

    df, n_corr, hoy = cargar_y_limpiar(io.BytesIO(raw.getvalue()), hoy)
    gen, ct, tab, meses = construir_tablas(df, hoy)

    wb = Workbook()
    hoja_procesos(wb, df, hoy, meses)
    hoja_general(wb, gen, hoy)
    hoja_mensual(wb, ct, meses)
    hoja_profesional(wb, tab, hoy)
    hoja_base_datos(wb, io.BytesIO(raw.getvalue()))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resumen = dict(
        evoluciones  = len(df),
        pacientes    = df['DOCUMENTO'].nunique(),
        profesionales= df['PROFESIONAL-ATIENDE'].nunique(),
        meses        = meses,
        correcciones = n_corr,
        hojas        = [s.title for s in wb.worksheets],
    )
    return output.getvalue(), resumen
